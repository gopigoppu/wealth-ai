from google.cloud import bigquery, storage
from typing import List, Dict
import os

import PyPDF2
from docx import Document
import openpyxl  # for .xlsx and .xls

BUCKET_NAME = "intelliwealth-processed-insights"
PREFIX = "thoughts/"


def extract_text_from_blob(blob) -> (str, str):
    name = blob.name.lower()
    try:
        if name.endswith(('.txt', '.md')):
            return blob.download_as_text(), ""
        elif name.endswith('.pdf'):
            reader = PyPDF2.PdfReader(blob.download_as_bytes())
            text = ''
            for page in reader.pages:
                text += page.extract_text() or ''
            return text, ""
        elif name.endswith('.docx'):
            data = blob.download_as_bytes()
            tmppath = "/tmp/temp.docx"
            with open(tmppath, "wb") as f:
                f.write(data)
            doc = Document(tmppath)
            return '\n'.join([p.text for p in doc.paragraphs]), ""
        elif name.endswith(('.xlsx', '.xls')):
            data = blob.download_as_bytes()
            tmppath = "/tmp/temp.xlsx"
            with open(tmppath, "wb") as f:
                f.write(data)
            wb = openpyxl.load_workbook(tmppath)
            text = ''
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    text += ' | '.join([str(cell)
                                       if cell is not None else '' for cell in row]) + '\n'
            return text, ""
        else:
            return "", f"Format not supported: {name}"
    except Exception as e:
        return "", f"Failed to extract {blob.name}: {str(e)}"


def vector_search_tool(query: str) -> List[Dict[str, str]]:
    results = []
    errors = []
    try:
        client = storage.Client()  # No credentials file, picks up Cloud Run's IAM identity!
        bucket = client.bucket(BUCKET_NAME)
        blobs = list(client.list_blobs(BUCKET_NAME, prefix=PREFIX))
        for blob in blobs:
            lower_name = blob.name.lower()
            if lower_name.endswith(('.txt', '.md', '.pdf', '.docx', '.xlsx', '.xls')):
                text, err = extract_text_from_blob(blob)
                if err:
                    errors.append(f"{blob.name}: {err}")
                    continue
                if query.lower() in text.lower():
                    idx = text.lower().find(query.lower())
                    snippet = text[max(0, idx-60):idx+len(query)+60]
                    results.append({
                        "summary": snippet.replace('\n', ' ').strip(),
                        "source": f"gs://{BUCKET_NAME}/{blob.name}"
                    })
        if not results:
            fallback_msg = f"No relevant doc found for '{query}'."
            if errors:
                fallback_msg += " (Some files could not be read: " + \
                    '; '.join(errors) + ")"
            results.append({
                "summary": fallback_msg,
                "source": f"gs://{BUCKET_NAME}/{PREFIX}"
            })
        elif errors:
            # Optionally, let user know some files were skipped
            results.append({
                "summary": f"Note: {len(errors)} file(s) could not be read. {', '.join(errors)}",
                "source": "error"
            })
        return results
    except Exception as e:
        return [{
            "summary": f"Sorry, the file search failed due to an internal error: {str(e)}",
            "source": "error"
        }]
